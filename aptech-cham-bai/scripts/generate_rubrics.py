import sys
import json
import openpyxl
import os
from openpyxl.styles import Border, Side, Alignment, Font

def main():
    if len(sys.argv) < 4:
        print("Usage: python generate_rubrics.py <data_json_path> <output_dir> <template_path>")
        sys.exit(1)
        
    data_path = sys.argv[1]
    out_dir = sys.argv[2]
    template_path = sys.argv[3]
    
    with open(data_path, 'r', encoding='utf-8') as f:
        data_list = json.load(f)
        
    os.makedirs(out_dir, exist_ok=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for item in data_list:
        fname = item["filename"]
        word_filename = item["word_filename"]
        reqs = item["reqs"]
        
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        word_filename_no_ext = os.path.splitext(word_filename)[0]
        ws.title = word_filename_no_ext[:31]
        
        # Replace "Tên môn thi" safely
        for r in range(1, 25):
            for c in range(1, 15):
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    if str(cell.value).strip() == "Tên môn thi":
                        cell.value = word_filename_no_ext

        needed_rows = len(reqs)
        available_rows = 15
        
        if needed_rows > available_rows:
            ws.insert_rows(18, amount=(needed_rows - available_rows))
            sum_row = 19 + (needed_rows - available_rows)
            formula_range = f"C4:C{sum_row-1}"
        else:
            sum_row = 19
            formula_range = f"C4:C18"

        # Clear columns A-G from row 4 to sum_row-1
        for r in range(4, sum_row):
            for c in range(1, 8):
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    cell.value = None
                    cell.border = thin_border
                    
        current_row = 4
        for idx, (req, score) in enumerate(reqs, start=1):
            c1 = ws.cell(row=current_row, column=1)
            c2 = ws.cell(row=current_row, column=2)
            c3 = ws.cell(row=current_row, column=3)
            c4 = ws.cell(row=current_row, column=4)
            c5 = ws.cell(row=current_row, column=5)
            c6 = ws.cell(row=current_row, column=6)
            
            c1.value = idx
            c2.value = req
            c3.value = float(score)
            
            for c in [c1, c2, c3, c4, c5, c6]:
                c.alignment = Alignment(wrap_text=True, vertical='top')
                
            current_row += 1
            
        # Update the Sum row
        c2_sum = ws.cell(row=sum_row, column=2)
        c3_sum = ws.cell(row=sum_row, column=3)
        c2_sum.value = "Sum:"
        c2_sum.font = Font(bold=True)
        c2_sum.alignment = Alignment(horizontal='right')
        
        c3_sum.value = f"=SUM({formula_range})"
        c3_sum.font = Font(bold=True)
        
        out_file = os.path.join(out_dir, fname)
        wb.save(out_file)
        print(f"Created {out_file} successfully.")

if __name__ == '__main__':
    main()
